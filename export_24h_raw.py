"""
Export all 1-minute targeted scrape data to Excel.

Produces one workbook per hub that has poll_interval_min=1 data:
  Sheet 1 – Snapshots   : per-minute hub-level charging counts
  Sheet 2 – EVSE Events : per-EVSE status log (every recorded status per poll)
  Sheet 3 – Visits      : completed + open sessions

Usage:
    python export_24h_raw.py                     # all hubs with 1-min data, all time
    python export_24h_raw.py --since 2026-04-27  # restrict to on/after this date (UTC)
"""

import argparse
import logging
import sqlite3
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import db

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("export_24h_raw")

_BLUE = PatternFill("solid", fgColor="4472C4")
_BLUE_FONT = Font(bold=True, color="FFFFFF")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_POLL = 1  # poll_interval_min — we only ever ran 1-min scrapes


def _hdr(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _BLUE
    c.font = _BLUE_FONT
    c.alignment = _CENTER
    return c


def _set_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_workbook(con: sqlite3.Connection, hub_uuid: str, hub_name: str, since: str | None) -> Workbook:
    since_clause = f"AND scraped_at >= '{since}'" if since else ""
    since_clause_started = f"AND started_at >= '{since}'" if since else ""

    wb = Workbook()

    # ── Sheet 1: Snapshots ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Snapshots"
    headers1 = [
        "Timestamp (UTC)", "Charging", "Available", "Inoperative",
        "Out of Order", "Unknown", "Total EVSEs", "Utilisation %",
    ]
    for col, h in enumerate(headers1, start=1):
        _hdr(ws1, 1, col, h)

    # Snapshots come from targeted_snapshots (counts) + snapshots (full counts) for same hub/time
    rows1 = con.execute(f"""
        SELECT
            ts.scraped_at,
            ts.charging_count,
            s.available_count,
            s.inoperative_count,
            s.out_of_order_count,
            s.unknown_count,
            h.total_evses,
            ts.utilisation_pct
        FROM targeted_snapshots ts
        JOIN hubs h ON h.uuid = ts.hub_uuid
        LEFT JOIN snapshots s
            ON s.hub_uuid = ts.hub_uuid
            AND substr(s.scraped_at, 1, 16) = substr(ts.scraped_at, 1, 16)
            AND s.source = 'targeted'
        WHERE ts.hub_uuid = ?
          AND ts.poll_interval_min = ?
          {since_clause}
        ORDER BY ts.scraped_at
    """, (hub_uuid, _POLL)).fetchall()

    for ri, r in enumerate(rows1, start=2):
        ws1.cell(row=ri, column=1, value=r[0])
        for col, val in enumerate(r[1:], start=2):
            ws1.cell(row=ri, column=col, value=val)

    ws1.freeze_panes = "A2"
    _set_widths(ws1, [22, 10, 10, 12, 13, 9, 12, 14])

    # ── Sheet 2: EVSE Events ────────────────────────────────────────────────
    ws2 = wb.create_sheet("EVSE Events")
    headers2 = ["Timestamp (UTC)", "EVSE UUID", "Status"]
    for col, h in enumerate(headers2, start=1):
        _hdr(ws2, 1, col, h)

    rows2 = con.execute(f"""
        SELECT scraped_at, evse_uuid, status
        FROM targeted_evse_events
        WHERE hub_uuid = ? AND poll_interval_min = ?
          {since_clause}
        ORDER BY scraped_at, evse_uuid
    """, (hub_uuid, _POLL)).fetchall()

    for ri, r in enumerate(rows2, start=2):
        ws2.cell(row=ri, column=1, value=r[0])
        ws2.cell(row=ri, column=2, value=r[1])
        ws2.cell(row=ri, column=3, value=r[2])

    ws2.freeze_panes = "A2"
    _set_widths(ws2, [22, 42, 14])

    # ── Sheet 3: Visits ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Visits")
    headers3 = ["EVSE UUID", "Started At (UTC)", "Ended At (UTC)", "Dwell (min)", "Status"]
    for col, h in enumerate(headers3, start=1):
        _hdr(ws3, 1, col, h)

    rows3 = con.execute(f"""
        SELECT evse_uuid, started_at, ended_at, dwell_min,
               CASE WHEN ended_at IS NULL THEN 'open' ELSE 'complete' END AS status
        FROM targeted_visits
        WHERE hub_uuid = ? AND poll_interval_min = ?
          {since_clause_started}
        ORDER BY started_at, evse_uuid
    """, (hub_uuid, _POLL)).fetchall()

    for ri, r in enumerate(rows3, start=2):
        for col, val in enumerate(r, start=1):
            ws3.cell(row=ri, column=col, value=val)

    ws3.freeze_panes = "A2"
    _set_widths(ws3, [42, 22, 22, 13, 10])

    # ── Summary row counts in each sheet title ───────────────────────────────
    ws1.title = f"Snapshots ({len(rows1)})"
    ws2.title = f"EVSE Events ({len(rows2)})"
    ws3.title = f"Visits ({len(rows3)})"

    return wb


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--since", help="ISO date or datetime, e.g. 2026-04-27")
    parser.add_argument("--out", default="exports", help="Output directory")
    args = parser.parse_args()

    out_dir = Path(args.out)
    out_dir.mkdir(exist_ok=True)

    since = args.since
    today_str = date.today().strftime("%Y-%m-%d")

    con = sqlite3.connect(db.DB_PATH)
    con.row_factory = sqlite3.Row

    hubs = con.execute("""
        SELECT DISTINCT h.uuid, h.hub_name
        FROM targeted_snapshots ts
        JOIN hubs h ON h.uuid = ts.hub_uuid
        WHERE ts.poll_interval_min = ?
        ORDER BY h.hub_name
    """, (_POLL,)).fetchall()

    if not hubs:
        log.warning("No targeted_snapshots rows found for poll_interval_min=%d", _POLL)
        con.close()
        return

    for hub in hubs:
        uuid = hub["uuid"]
        name = hub["hub_name"] or uuid
        log.info("Exporting '%s' ...", name)

        wb = _build_workbook(con, uuid, name, since)

        slug = name.lower().replace(" ", "_").replace("/", "_")
        fname = f"raw_1min_{slug}_{today_str}.xlsx"
        path = out_dir / fname
        wb.save(path)
        log.info("  Saved → %s", path)

    con.close()


if __name__ == "__main__":
    main()
