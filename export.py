"""Daily Excel report generator — all sites and per-group exports."""
import logging
import re
import sqlite3
from calendar import monthrange
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter

import db

log = logging.getLogger("evanti.export")

_PCT_FMT = "0.00%"
_PCT_STYLE = "pct"
_BLUE = PatternFill("solid", fgColor="4472C4")
_BLUE_FONT = Font(bold=True, color="FFFFFF")
_LIGHT_BLUE = PatternFill("solid", fgColor="8EAADB")
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _slug(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def _months_between(start: str, end: str) -> list[str]:
    """Return ['YYYY-MM', ...] from start to end inclusive."""
    sy, sm = int(start[:4]), int(start[5:7])
    ey, em = int(end[:4]), int(end[5:7])
    result, y, m = [], sy, sm
    while (y, m) <= (ey, em):
        result.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return result


def _month_label(ym: str) -> str:
    """'2026-04' → 'Apr-26'"""
    return date(int(ym[:4]), int(ym[5:7]), 1).strftime("%b-%y")


def _query_export_data(hub_uuids: list[str] | None) -> dict:
    con = sqlite3.connect(db.DB_PATH)
    con.row_factory = sqlite3.Row

    today = date.today()
    cur_month = today.strftime("%Y-%m")
    p = list(hub_uuids) if hub_uuids else []
    ph = ",".join("?" * len(p))

    def q(sql_all: str, sql_filtered: str, extra: tuple = ()) -> list:
        """Run sql_filtered with (extra + p) if filtering by uuids, else sql_all with extra."""
        if p:
            return con.execute(sql_filtered, list(extra) + p).fetchall()
        return con.execute(sql_all, list(extra)).fetchall()

    hubs = q(
        "SELECT uuid, hub_name, operator, max_power_kw, total_evses FROM hubs ORDER BY hub_name",
        f"SELECT uuid, hub_name, operator, max_power_kw, total_evses FROM hubs WHERE uuid IN ({ph}) ORDER BY hub_name",
    )

    totals_util = {r["hub_uuid"]: r["v"] for r in q(
        "SELECT hub_uuid, AVG(utilisation_pct) / 100.0 AS v FROM snapshots GROUP BY hub_uuid",
        f"SELECT hub_uuid, AVG(utilisation_pct) / 100.0 AS v FROM snapshots WHERE hub_uuid IN ({ph}) GROUP BY hub_uuid",
    )}

    totals_visits = {r["hub_uuid"]: r["v"] for r in q(
        "SELECT hub_uuid, COUNT(*) AS v FROM visits GROUP BY hub_uuid",
        f"SELECT hub_uuid, COUNT(*) AS v FROM visits WHERE hub_uuid IN ({ph}) GROUP BY hub_uuid",
    )}

    monthly_util: dict[str, dict[str, float]] = {}
    for r in q(
        "SELECT hub_uuid, substr(scraped_at,1,7) AS mo, AVG(utilisation_pct) / 100.0 AS v FROM snapshots GROUP BY hub_uuid, mo",
        f"SELECT hub_uuid, substr(scraped_at,1,7) AS mo, AVG(utilisation_pct) / 100.0 AS v FROM snapshots WHERE hub_uuid IN ({ph}) GROUP BY hub_uuid, mo",
    ):
        monthly_util.setdefault(r["hub_uuid"], {})[r["mo"]] = r["v"]

    monthly_visits: dict[str, dict[str, int]] = {}
    for r in q(
        "SELECT hub_uuid, substr(started_at,1,7) AS mo, COUNT(*) AS v FROM visits GROUP BY hub_uuid, mo",
        f"SELECT hub_uuid, substr(started_at,1,7) AS mo, COUNT(*) AS v FROM visits WHERE hub_uuid IN ({ph}) GROUP BY hub_uuid, mo",
    ):
        monthly_visits.setdefault(r["hub_uuid"], {})[r["mo"]] = r["v"]

    daily_util: dict[str, dict[str, float]] = {}
    for r in q(
        "SELECT hub_uuid, substr(scraped_at,1,10) AS dy, AVG(utilisation_pct) / 100.0 AS v FROM snapshots WHERE substr(scraped_at,1,7)=? GROUP BY hub_uuid, dy",
        f"SELECT hub_uuid, substr(scraped_at,1,10) AS dy, AVG(utilisation_pct) / 100.0 AS v FROM snapshots WHERE substr(scraped_at,1,7)=? AND hub_uuid IN ({ph}) GROUP BY hub_uuid, dy",
        extra=(cur_month,),
    ):
        daily_util.setdefault(r["hub_uuid"], {})[r["dy"]] = r["v"]

    daily_visits: dict[str, dict[str, int]] = {}
    for r in q(
        "SELECT hub_uuid, substr(started_at,1,10) AS dy, COUNT(*) AS v FROM visits WHERE substr(started_at,1,7)=? GROUP BY hub_uuid, dy",
        f"SELECT hub_uuid, substr(started_at,1,10) AS dy, COUNT(*) AS v FROM visits WHERE substr(started_at,1,7)=? AND hub_uuid IN ({ph}) GROUP BY hub_uuid, dy",
        extra=(cur_month,),
    ):
        daily_visits.setdefault(r["hub_uuid"], {})[r["dy"]] = r["v"]

    con.close()

    all_months: set[str] = set()
    for d in (monthly_util, monthly_visits):
        for hub_months in d.values():
            all_months.update(hub_months.keys())
    months = _months_between(min(all_months), cur_month) if all_months else [cur_month]

    _, days_in_month = monthrange(today.year, today.month)
    days = [f"{today.year:04d}-{today.month:02d}-{d:02d}" for d in range(1, days_in_month + 1)]

    return {
        "hubs": [dict(r) for r in hubs],
        "totals_util": totals_util,
        "totals_visits": totals_visits,
        "monthly_util": monthly_util,
        "monthly_visits": monthly_visits,
        "daily_util": daily_util,
        "daily_visits": daily_visits,
        "months": months,
        "days": days,
    }


def _sh(ws, row: int, col: int, value, fill=None, font=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill or _BLUE
    c.font = font or _BLUE_FONT
    c.alignment = _CENTER


def _build_workbook(data: dict) -> Workbook:
    hubs = data["hubs"]
    months = data["months"]
    days = data["days"]
    N, D = len(months), len(days)

    C_HUB, C_TOWN, C_OP, C_DESC, C_KW, C_EVSES, C_TU, C_TV = 1, 2, 3, 4, 5, 6, 7, 8
    C_MU = 9
    C_MV = C_MU + N
    C_DU = C_MV + N
    C_DV = C_DU + D

    wb = Workbook()
    pct = NamedStyle(name=_PCT_STYLE, number_format=_PCT_FMT)
    wb.add_named_style(pct)
    ws = wb.active
    ws.title = "Report"

    # Row 1: super-headers
    for col in range(1, 9):
        _sh(ws, 1, col, None)
    if N:
        ws.merge_cells(start_row=1, start_column=C_MU, end_row=1, end_column=C_MU + N - 1)
        _sh(ws, 1, C_MU, "Average Utilisation (%)")
        ws.merge_cells(start_row=1, start_column=C_MV, end_row=1, end_column=C_MV + N - 1)
        _sh(ws, 1, C_MV, "Total Number of Visits")
    if D:
        ws.merge_cells(start_row=1, start_column=C_DU, end_row=1, end_column=C_DU + D - 1)
        _sh(ws, 1, C_DU, "Daily Utilisation (%)")
        ws.merge_cells(start_row=1, start_column=C_DV, end_row=1, end_column=C_DV + D - 1)
        _sh(ws, 1, C_DV, "Daily Number of Visits")

    # Row 2: sub-headers
    for col, label in [
        (C_HUB, "Hub"), (C_TOWN, "Town"), (C_OP, "Operator"), (C_DESC, "Description"),
        (C_KW, "Max kW"), (C_EVSES, "EVSEs"),
        (C_TU, "Total Average Utilisation (%)"), (C_TV, "Total Number of Visits"),
    ]:
        _sh(ws, 2, col, label, fill=_LIGHT_BLUE, font=_BOLD)

    for i, ym in enumerate(months):
        lbl = _month_label(ym)
        _sh(ws, 2, C_MU + i, lbl, fill=_LIGHT_BLUE, font=_BOLD)
        _sh(ws, 2, C_MV + i, lbl, fill=_LIGHT_BLUE, font=_BOLD)

    today = date.today()
    for i, day_str in enumerate(days):
        d = date(today.year, today.month, int(day_str[8:10]))
        for base in (C_DU, C_DV):
            c = ws.cell(row=2, column=base + i, value=d)
            c.fill = _LIGHT_BLUE
            c.font = _BOLD
            c.alignment = _CENTER
            c.number_format = "D-MMM"

    # Data rows
    for ri, hub in enumerate(hubs, start=3):
        uuid = hub["uuid"]

        ws.cell(row=ri, column=C_HUB, value=hub.get("hub_name") or "")
        ws.cell(row=ri, column=C_TOWN, value=None)
        ws.cell(row=ri, column=C_OP, value=hub.get("operator") or "")
        ws.cell(row=ri, column=C_DESC, value=None)
        ws.cell(row=ri, column=C_KW, value=hub.get("max_power_kw"))
        ws.cell(row=ri, column=C_EVSES, value=hub.get("total_evses"))

        tu = data["totals_util"].get(uuid)
        c = ws.cell(row=ri, column=C_TU, value=tu)
        if tu is not None:
            c.style = _PCT_STYLE

        ws.cell(row=ri, column=C_TV, value=data["totals_visits"].get(uuid) or 0)

        hub_mu = data["monthly_util"].get(uuid, {})
        for i, ym in enumerate(months):
            v = hub_mu.get(ym)
            c = ws.cell(row=ri, column=C_MU + i, value=v)
            if v is not None:
                c.style = _PCT_STYLE

        hub_mv = data["monthly_visits"].get(uuid, {})
        for i, ym in enumerate(months):
            ws.cell(row=ri, column=C_MV + i, value=hub_mv.get(ym))

        hub_du = data["daily_util"].get(uuid, {})
        for i, day_str in enumerate(days):
            v = hub_du.get(day_str)
            c = ws.cell(row=ri, column=C_DU + i, value=v)
            if v is not None:
                c.style = _PCT_STYLE

        hub_dv = data["daily_visits"].get(uuid, {})
        for i, day_str in enumerate(days):
            ws.cell(row=ri, column=C_DV + i, value=hub_dv.get(day_str))

    # Column widths
    ws.column_dimensions[get_column_letter(C_HUB)].width = 35
    ws.column_dimensions[get_column_letter(C_TOWN)].width = 15
    ws.column_dimensions[get_column_letter(C_OP)].width = 25
    ws.column_dimensions[get_column_letter(C_DESC)].width = 30
    ws.column_dimensions[get_column_letter(C_KW)].width = 10
    ws.column_dimensions[get_column_letter(C_EVSES)].width = 8
    ws.column_dimensions[get_column_letter(C_TU)].width = 22
    ws.column_dimensions[get_column_letter(C_TV)].width = 20
    for i in range(N):
        ws.column_dimensions[get_column_letter(C_MU + i)].width = 9
        ws.column_dimensions[get_column_letter(C_MV + i)].width = 9
    for i in range(D):
        ws.column_dimensions[get_column_letter(C_DU + i)].width = 7
        ws.column_dimensions[get_column_letter(C_DV + i)].width = 7

    ws.freeze_panes = "C3"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 45

    return wb



def _build_interval_workbook(
    hub_name: str,
    intervals: list[int],
    snap_data: dict,   # {interval: [{scraped_at, charging_count, utilisation_pct}]}
    visit_data: dict,  # {interval: {total_visits, visit_days, avg_dwell_min}}
) -> Workbook:
    """Three-sheet workbook with real per-interval data from targeted_* tables."""
    wb = Workbook()
    pct = NamedStyle(name=_PCT_STYLE, number_format=_PCT_FMT)
    wb.add_named_style(pct)

    # ── Sheet 1: Timeline ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Timeline"
    _sh(ws1, 1, 1, "Timestamp")
    for col, iv in enumerate(intervals, start=2):
        _sh(ws1, 1, col, f"Charging ({iv}m)")

    snaps_by_interval: dict[int, dict[str, int]] = {}
    for iv in intervals:
        for s in snap_data.get(iv, []):
            ts = s["scraped_at"][:16]
            snaps_by_interval.setdefault(iv, {})[ts] = s["charging_count"]

    all_timestamps = sorted({ts for iv in intervals for ts in snaps_by_interval.get(iv, {})})
    for ri, ts in enumerate(all_timestamps, start=2):
        ws1.cell(row=ri, column=1, value=ts)
        for col, iv in enumerate(intervals, start=2):
            v = snaps_by_interval.get(iv, {}).get(ts)
            ws1.cell(row=ri, column=col, value=v if v is not None else "")

    ws1.freeze_panes = "B2"
    ws1.column_dimensions["A"].width = 18
    for i in range(len(intervals)):
        ws1.column_dimensions[get_column_letter(2 + i)].width = 14

    # ── Sheet 2: Visits ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Visits")
    headers2 = ["Interval", "Total Visits", "Visit Days", "Visits/Day", "Avg Dwell (min)"]
    for col, h in enumerate(headers2, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = _BLUE_FONT
        c.fill = _BLUE

    for ri, iv in enumerate(intervals, start=2):
        stats = visit_data.get(iv, {})
        tv = stats.get("total_visits", 0)
        vd = stats.get("visit_days") or 0
        ws2.cell(row=ri, column=1, value=f"{iv} min")
        ws2.cell(row=ri, column=2, value=tv)
        ws2.cell(row=ri, column=3, value=vd)
        ws2.cell(row=ri, column=4, value=round(tv / vd, 1) if vd else 0)
        ws2.cell(row=ri, column=5, value=stats.get("avg_dwell_min"))

    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # ── Sheet 3: Miss Rate ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Miss Rate")
    baseline = visit_data.get(intervals[0], {}).get("total_visits") or 0
    _sh(ws3, 1, 1, "Interval")
    _sh(ws3, 1, 2, "Visits Detected")
    _sh(ws3, 1, 3, f"vs {intervals[0]}-min baseline")
    _sh(ws3, 1, 4, "Miss Rate")

    for ri, iv in enumerate(intervals, start=2):
        detected = visit_data.get(iv, {}).get("total_visits", 0)
        missed = baseline - detected
        ws3.cell(row=ri, column=1, value=f"{iv} min")
        ws3.cell(row=ri, column=2, value=detected)
        ws3.cell(row=ri, column=3, value=missed)
        c = ws3.cell(row=ri, column=4, value=(missed / baseline) if baseline else 0)
        c.number_format = "0.0%"

    for col in range(1, 5):
        ws3.column_dimensions[get_column_letter(col)].width = 20

    return wb


def export_interval_comparison(output_dir: str | Path = "exports") -> list[Path]:
    """Per-hub workbooks comparing real independent per-interval polling data."""
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)
    today_str = date.today().strftime("%Y-%m-%d")
    window_start = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    generated: list[Path] = []

    con = sqlite3.connect(db.DB_PATH)
    con.row_factory = sqlite3.Row

    hub_rows = con.execute("""
        SELECT DISTINCT h.uuid, h.hub_name
        FROM hubs h
        JOIN group_hubs gh ON gh.hub_uuid = h.uuid
        JOIN groups g ON g.id = gh.group_id
        WHERE g.scrape_interval IS NOT NULL
        ORDER BY h.hub_name
    """).fetchall()

    for hub in hub_rows:
        uuid, name = hub["uuid"], hub["hub_name"] or hub["uuid"]

        interval_rows = con.execute("""
            SELECT DISTINCT g.scrape_interval
            FROM groups g
            JOIN group_hubs gh ON gh.group_id = g.id
            WHERE gh.hub_uuid = ? AND g.scrape_interval IS NOT NULL
            ORDER BY g.scrape_interval
        """, (uuid,)).fetchall()
        intervals = [r["scrape_interval"] for r in interval_rows]
        if not intervals:
            continue

        # Real per-interval snapshots from targeted_snapshots
        snap_data: dict[int, list] = {}
        for iv in intervals:
            rows = con.execute("""
                SELECT scraped_at, charging_count, utilisation_pct
                FROM targeted_snapshots
                WHERE hub_uuid = ? AND poll_interval_min = ? AND scraped_at >= ?
                ORDER BY scraped_at
            """, (uuid, iv, window_start)).fetchall()
            snap_data[iv] = [dict(r) for r in rows]

        if not any(snap_data.values()):
            continue

        # Real per-interval visit stats from targeted_visits
        visit_data: dict[int, dict] = {}
        for iv in intervals:
            row = con.execute("""
                SELECT COUNT(*) AS total_visits,
                       COUNT(DISTINCT DATE(started_at)) AS visit_days,
                       ROUND(AVG(CASE WHEN ended_at IS NOT NULL THEN dwell_min END)) AS avg_dwell_min
                FROM targeted_visits
                WHERE hub_uuid = ? AND poll_interval_min = ? AND started_at >= ?
            """, (uuid, iv, window_start)).fetchone()
            visit_data[iv] = dict(row) if row else {}

        wb = _build_interval_workbook(name, intervals, snap_data, visit_data)
        path = output_dir / f"interval_comparison_{_slug(name)}_{today_str}.xlsx"
        wb.save(path)
        log.info("Saved %s (intervals %s)", path, intervals)
        generated.append(path)

    con.close()
    return generated


def export_reports(output_dir: str | Path = "exports") -> list[Path]:
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)
    month_str = date.today().strftime("%Y-%m")
    generated: list[Path] = []

    log.info("Generating all-sites export...")
    data = _query_export_data(None)
    wb = _build_workbook(data)
    path = output_dir / f"all_sites_{month_str}.xlsx"
    wb.save(path)
    log.info("Saved %s (%d hubs)", path, len(data["hubs"]))
    generated.append(path)

    con = sqlite3.connect(db.DB_PATH)
    con.row_factory = sqlite3.Row
    groups = [dict(r) for r in con.execute("SELECT id, name FROM groups ORDER BY id").fetchall()]
    group_hubs = {}
    for r in con.execute("SELECT group_id, hub_uuid FROM group_hubs").fetchall():
        group_hubs.setdefault(r["group_id"], []).append(r["hub_uuid"])
    con.close()

    for group in groups:
        uuids = group_hubs.get(group["id"], [])
        if not uuids:
            continue
        log.info("Generating export for group '%s' (%d hubs)...", group["name"], len(uuids))
        data = _query_export_data(uuids)
        wb = _build_workbook(data)
        path = output_dir / f"{_slug(group['name'])}_{month_str}.xlsx"
        wb.save(path)
        log.info("Saved %s", path)
        generated.append(path)

    return generated


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    for p in export_reports():
        print(p)
    for p in export_interval_comparison():
        print(p)
